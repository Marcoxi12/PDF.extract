import streamlit as st
import io
import re
import pdfplumber
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="FAE Invoice Processor",
    page_icon="🔧",
    layout="centered",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'IBM Plex Sans', sans-serif;
    }
    .main { background: #0d0d0d; }
    .block-container { max-width: 760px; padding-top: 2rem; }

    .hero {
        background: #111;
        border: 1px solid #222;
        border-radius: 4px;
        padding: 2.5rem 2rem 2rem;
        margin-bottom: 2rem;
        text-align: center;
        position: relative;
        overflow: hidden;
    }
    .hero::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, #ff6b00, #ffaa00, #ff6b00);
    }
    .hero h1 {
        color: #f0f0f0;
        font-size: 1.75rem;
        font-weight: 600;
        margin: 0 0 0.4rem 0;
        font-family: 'IBM Plex Mono', monospace;
        letter-spacing: -1px;
    }
    .hero p { color: #666; font-size: 0.9rem; margin: 0; }
    .hero .icon { font-size: 2.2rem; margin-bottom: 0.75rem; display: block; }

    .step-card {
        background: #111;
        border-radius: 4px;
        padding: 1.1rem 1.4rem;
        margin-bottom: 0.6rem;
        border: 1px solid #222;
        color: #aaa;
        font-size: 0.9rem;
    }
    .step-num {
        display: inline-block;
        background: #ff6b00;
        color: white;
        border-radius: 2px;
        width: 22px; height: 22px;
        text-align: center;
        line-height: 22px;
        font-size: 0.75rem;
        font-weight: 700;
        margin-right: 0.6rem;
        font-family: 'IBM Plex Mono', monospace;
    }

    .stDownloadButton > button {
        background: #ff6b00 !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        padding: 0.7rem 2rem !important;
        font-size: 0.95rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        margin-top: 0.5rem !important;
        font-family: 'IBM Plex Mono', monospace !important;
        letter-spacing: 0.5px !important;
        transition: all 0.15s !important;
    }
    .stDownloadButton > button:hover {
        background: #e55c00 !important;
        transform: translateY(-1px) !important;
    }

    .success-box {
        background: #0a1a0a;
        border: 1px solid #1a4a1a;
        border-left: 3px solid #2a8a2a;
        border-radius: 4px;
        padding: 0.85rem 1.1rem;
        margin: 1rem 0;
        color: #6abf6a;
        font-size: 0.88rem;
        font-family: 'IBM Plex Mono', monospace;
    }
    .warn-box {
        background: #1a1100;
        border: 1px solid #3a2800;
        border-left: 3px solid #ff6b00;
        border-radius: 4px;
        padding: 0.85rem 1.1rem;
        margin: 1rem 0;
        color: #cc8844;
        font-size: 0.88rem;
    }
    .stat-row {
        display: flex;
        gap: 0.75rem;
        margin: 1rem 0;
    }
    .stat {
        flex: 1;
        background: #111;
        border-radius: 4px;
        padding: 1rem 0.75rem;
        text-align: center;
        border: 1px solid #222;
    }
    .stat-val {
        font-size: 1.5rem;
        font-weight: 700;
        color: #ff6b00;
        font-family: 'IBM Plex Mono', monospace;
    }
    .stat-lbl { font-size: 0.72rem; color: #555; margin-top: 3px; text-transform: uppercase; letter-spacing: 0.5px; }
    .footer {
        text-align: center;
        color: #333;
        font-size: 0.75rem;
        margin-top: 2rem;
        padding-top: 1rem;
        border-top: 1px solid #1a1a1a;
        font-family: 'IBM Plex Mono', monospace;
    }
    .stFileUploader { border-radius: 4px !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def parse_money(s):
    if s is None:
        return None
    cleaned = str(s).replace("$", "").replace(",", "").replace(" ", "").strip()
    neg = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    try:
        val = float(cleaned) if cleaned else None
    except ValueError:
        return None
    if val is None:
        return None
    return -val if neg else val

def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    b = Side(style="medium", color="595959")
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=b)

MONEY_FMT    = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
DASH_FMT     = '_($* #,##0.00_);_($* (#,##0.00);"-";_(@_)'
HDR_FILL     = PatternFill("solid", start_color="1F1F1F")
HDR_FONT     = Font(color="FFFFFF", bold=True, name="Arial", size=9)
DATA_FONT    = Font(name="Arial", size=9)
BOLD_FONT    = Font(name="Arial", size=9, bold=True)
YLW_FILL     = PatternFill("solid", start_color="FFFF00")
GRAY_FILL    = PatternFill("solid", start_color="F2F2F2")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
LTBLUE_FILL  = PatternFill("solid", start_color="DCE6F1")


# ─────────────────────────────────────────────
# PDF Parsing
# ─────────────────────────────────────────────

def parse_invoice(pdf_bytes):
    """
    Parse the FAE / Small Compressor invoice PDF.
    Returns a dict with invoice metadata and line items.
    """
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

        # Extract header metadata
        invoice_no = re.search(r"INVOICE NO\.?:?\s*(\S+)", text, re.I)
        date_m     = re.search(r"DATE:\s*(\S+)", text, re.I)
        months_pat = (r"\b(January|February|March|April|May|June|July|August"
                      r"|September|October|November|December)\b")
        month_m    = re.search(months_pat, text, re.I)

        invoice_number = invoice_no.group(1) if invoice_no else ""
        invoice_date   = date_m.group(1) if date_m else ""
        for_month      = month_m.group(1).capitalize() if month_m else ""

        # Get subtotal / sales tax / total
        sub_m   = re.search(r"SUBTOTAL\s+\$?\s*([\d,]+\.\d{2})", text, re.I)
        tax_m   = re.search(r"SALES TAX\s+([\d,]+\.\d{2})", text, re.I)
        total_m = re.search(r"\bTOTAL\b\s+\$\s*([\d,]+\.\d{2})", text, re.I)

        subtotal   = parse_money(sub_m.group(1))   if sub_m   else 0.0
        sales_tax  = parse_money(tax_m.group(1))   if tax_m   else 0.0
        total      = parse_money(total_m.group(1)) if total_m else 0.0

        # Parse line items from table
        items = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 6:
                        continue
                    qty_raw  = row[0]
                    desc_raw = row[1]
                    up_raw   = row[4]
                    del_raw  = row[5]
                    add_raw  = row[6] if len(row) > 6 else None
                    lt_raw   = row[7] if len(row) > 7 else None

                    # Skip header rows, empty rows, summary rows
                    if not qty_raw or not desc_raw:
                        continue
                    qty_clean = str(qty_raw).strip().replace(" ", "")
                    try:
                        qty = float(qty_clean)
                    except ValueError:
                        continue
                    if qty == 0:
                        continue

                    desc       = str(desc_raw).strip() if desc_raw else ""
                    unit_price = parse_money(up_raw)
                    delivery   = parse_money(del_raw)
                    add_chg    = parse_money(add_raw)
                    line_total = parse_money(lt_raw)

                    # Split description into Name | Vendor Cross Reference
                    parts = desc.split("|", 1)
                    name  = parts[0].strip() if parts else desc
                    vendor_ref = parts[1].strip() if len(parts) > 1 else desc

                    items.append({
                        "qty":         qty,
                        "name":        name,
                        "vendor_ref":  vendor_ref,
                        "unit_price":  unit_price if unit_price else 0.0,
                        "delivery":    delivery if delivery else 0.0,
                        "add_charges": add_chg if add_chg else 0.0,
                        "line_total":  line_total if line_total else 0.0,
                    })

    return {
        "invoice_number": invoice_number,
        "invoice_date":   invoice_date,
        "for_month":      for_month,
        "subtotal":       subtotal,
        "sales_tax":      sales_tax,
        "total":          total,
        "items":          items,
    }


# ─────────────────────────────────────────────
# Excel Export  — matches the screenshot layout
# ─────────────────────────────────────────────

def build_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ── Column widths (matching screenshot) ──────────────────────────────────
    ws.column_dimensions["A"].width = 28   # Name
    ws.column_dimensions["B"].width = 62   # Vendor Cross Reference
    ws.column_dimensions["C"].width = 14   # UNIT PRICE
    ws.column_dimensions["D"].width = 14   # Delivery Fee
    ws.column_dimensions["E"].width = 14   # Add'l Charges
    ws.column_dimensions["F"].width = 14   # LINE TOTAL

    # ── Row 1: Invoice metadata header bar ───────────────────────────────────
    ws.row_dimensions[1].height = 18
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = (
        f"FAE II  ·  INVOICE NO.: {data['invoice_number']}  ·  "
        f"DATE: {data['invoice_date']}  ·  FOR THE MONTH OF: {data['for_month']}"
    )
    c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="1F1F1F")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: Column headers (matching screenshot) ───────────────────────────
    ws.row_dimensions[2].height = 30
    headers = ["Name", "Vendor Cross Refrence", "UNIT PRICE", "Delivery Fee", "Add'l Charges", "LINE TOTAL"]
    for ci, h in enumerate(headers, 1):
        c           = ws.cell(2, ci, h)
        c.font      = Font(name="Arial", size=9, bold=True)
        c.fill      = PatternFill("solid", start_color="EDEDED")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thick_bottom()

    # ── Data rows ─────────────────────────────────────────────────────────────
    MONEY  = '$#,##0.00_);($#,##0.00)'
    DASH   = '_($* #,##0.00_);_($* (#,##0.00);"-";_(@_)'

    START_ROW = 3
    items = data["items"]

    for ri, item in enumerate(items, START_ROW):
        ws.row_dimensions[ri].height = 15

        # Alternate fill: white / very-light-blue like screenshot
        fill = WHITE_FILL if (ri - START_ROW) % 2 == 0 else LTBLUE_FILL

        # A: Name
        c = ws.cell(ri, 1, item["name"])
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin()
        c.fill      = fill

        # B: Vendor Cross Reference
        c = ws.cell(ri, 2, item["vendor_ref"])
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin()
        c.fill      = fill

        # C: Unit Price
        c = ws.cell(ri, 3, item["unit_price"])
        c.number_format = MONEY
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = thin()
        c.fill      = fill

        # D: Delivery Fee  (show "-" for zero)
        c = ws.cell(ri, 4, item["delivery"] if item["delivery"] != 0 else None)
        c.number_format = DASH
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = thin()
        c.fill      = fill
        if item["delivery"] == 0:
            c.value = "-"
            c.alignment = Alignment(horizontal="center", vertical="center")

        # E: Add'l Charges (show "-" for zero, otherwise parentheses for negatives)
        add = item["add_charges"]
        c = ws.cell(ri, 5)
        if add == 0:
            c.value     = "-"
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.value         = add
            c.number_format = MONEY
            c.alignment     = Alignment(horizontal="right", vertical="center")
        c.font   = DATA_FONT
        c.border = thin()
        c.fill   = fill

        # F: Line Total  (show "-" for zero)
        lt = item["line_total"]
        c = ws.cell(ri, 6)
        if lt == 0:
            c.value     = "-"
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.value         = lt
            c.number_format = MONEY
            c.alignment     = Alignment(horizontal="right", vertical="center")
        c.font   = DATA_FONT
        c.border = thin()
        c.fill   = fill

    # ── Totals section ────────────────────────────────────────────────────────
    last_data = START_ROW + len(items) - 1
    gap_row   = last_data + 1
    sub_row   = last_data + 2
    tax_row   = last_data + 3
    tot_row   = last_data + 4

    # Blank spacer row
    ws.row_dimensions[gap_row].height = 6

    def totals_row(row_num, label, value, bold=False, fill_color="F9F9F9", border_top=False):
        ws.row_dimensions[row_num].height = 16
        fill = PatternFill("solid", start_color=fill_color)
        font = Font(name="Arial", size=9, bold=bold)

        # Merge A:E for label area
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        c = ws.cell(row_num, 1, label)
        c.font      = font
        c.fill      = fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        side = Side(style="thin", color="BFBFBF")
        top  = Side(style="medium" if border_top else "thin", color="595959" if border_top else "BFBFBF")
        c.border = Border(left=side, right=side, top=top, bottom=side)

        # Value in F
        c2 = ws.cell(row_num, 6, value)
        c2.number_format = '$#,##0.00_);($#,##0.00)'
        c2.font      = font
        c2.fill      = fill
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.border    = Border(left=side, right=side, top=top, bottom=side)

    totals_row(sub_row, "SUBTOTAL", data["subtotal"])
    totals_row(tax_row, "SALES TAX", data["sales_tax"])
    totals_row(tot_row, "TOTAL", data["total"], bold=True, fill_color="EDEDED", border_top=True)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <span class="icon">🔧</span>
    <h1>FAE Invoice Processor</h1>
    <p>Upload a Small Compressor Sales & Rentals invoice PDF → download a formatted Excel file instantly</p>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Drop your invoice PDF here",
    type=["pdf"],
    label_visibility="collapsed"
)

if not uploaded:
    st.markdown("""
    <div class="step-card"><span class="step-num">1</span> <strong>Upload</strong> a FAE / Small Compressor Sales invoice PDF above</div>
    <div class="step-card"><span class="step-num">2</span> <strong>Review</strong> the extracted line items and totals summary</div>
    <div class="step-card"><span class="step-num">3</span> <strong>Download</strong> a clean, formatted Excel file matching your invoice layout</div>
    """, unsafe_allow_html=True)
else:
    with st.spinner("Parsing invoice…"):
        try:
            pdf_bytes = uploaded.read()
            data      = parse_invoice(pdf_bytes)
            items     = data["items"]

            if not items:
                st.markdown('<div class="warn-box">⚠️ No line items found. Make sure this is a FAE / Small Compressor Sales invoice.</div>', unsafe_allow_html=True)
            else:
                excel_bytes = build_excel(data)
                filename    = uploaded.name.replace(".pdf", "").replace(".PDF", "")

                # Count stats
                non_zero   = [i for i in items if i["line_total"] != 0]
                zero_items = [i for i in items if i["line_total"] == 0]

                st.markdown(
                    f'<div class="success-box">✓ Parsed {len(items)} line items from invoice {data["invoice_number"]}</div>',
                    unsafe_allow_html=True
                )

                st.markdown(f"""
                <div class="stat-row">
                    <div class="stat">
                        <div class="stat-val">{len(items)}</div>
                        <div class="stat-lbl">Line Items</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">{len(zero_items)}</div>
                        <div class="stat-lbl">Zero-Total Lines</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">${data['subtotal']:,.2f}</div>
                        <div class="stat-lbl">Subtotal</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">${data['total']:,.2f}</div>
                        <div class="stat-lbl">Invoice Total</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.download_button(
                    label="⬇  Download Excel File",
                    data=excel_bytes,
                    file_name=f"{filename}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            import traceback
            st.markdown(
                f'<div class="warn-box">❌ Error processing file: {str(e)}<br><pre>{traceback.format_exc()}</pre></div>',
                unsafe_allow_html=True
            )

st.markdown('<div class="footer">FAE Invoice Processor · Small Compressor Sales & Rentals</div>', unsafe_allow_html=True)
